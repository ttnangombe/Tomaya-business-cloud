
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def export_excel(path: str, sheet: str, df: pd.DataFrame, title: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    start = 1
    if title:
        ws["A1"] = title
        start = 3
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
